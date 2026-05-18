#!/usr/bin/env python3
"""Append KEYBOARD-MULTI-IOS test cases to docs/test-cases.xlsx — 56 rows.

Mirrors the Android KEYBOARD-MULTI sheet adapted for iOS specifics:
- Swipe-space-bar cycles languages (no chip strip on iOS either, per
  the 2026-05-18 evening plan revision after Android shipping)
- System globe key (iOS's bottom-left 🌐) handles system keyboard
  switching — DraftRight does NOT intercept it
- Allow Full Access toggle + memory ceiling iOS-specific cases

Idempotent guard: aborts if the sheet already exists."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX = Path(__file__).resolve().parents[1] / "docs" / "test-cases.xlsx"
SHEET = "KEYBOARD-MULTI-IOS"
HEADERS = ["TC-ID", "Title", "Preconditions", "Steps", "Expected", "Verified by", "Priority", "Owner"]


def main() -> int:
    wb = load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        print(f"Sheet {SHEET} exists — aborting")
        return 1
    ws = wb.create_sheet(SHEET)
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    rows = [
        # ── Telex composition — core rules (12) ─────────────────────────
        ("KEYBOARD-MULTI-IOS-001", "Telex aa → â", "VI active", "Type a,a", "Composer commits â", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-002", "Telex oo → ô", "VI active", "Type o,o", "Composer commits ô", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-003", "Telex ee → ê", "VI active", "Type e,e", "Composer commits ê", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-004", "Telex ow → ơ", "VI active", "Type o,w", "Composer commits ơ", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-005", "Telex uw → ư", "VI active", "Type u,w", "Composer commits ư", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-006", "Telex aw → ă", "VI active", "Type a,w", "Composer commits ă", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-007", "Telex dd → đ", "VI active", "Type d,d", "Composer commits đ", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-008", "Telex aaj → ậ", "VI active", "Type a,a,j", "Composer commits ậ (â + dot)", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-009", "Telex uow → ươ", "VI active", "Type u,o,w", "Composer commits ươ", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-010", "Telex uowj → ượ", "VI active", "Type u,o,w,j", "Composer commits ượ", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-011", "Telex case preserved on shift", "VI active, shift on", "Type A,A", "Composer commits Â", "TelexComposerTests.swift", "P1", "Tan"),
        ("KEYBOARD-MULTI-IOS-012", "Telex no-op for non-rule keys", "VI active", "Type q", "Direct insertText q (no marked text)", "TelexComposerTests.swift", "P0", "Tan"),

        # ── Telex composition — tones (6) ────────────────────────────────
        ("KEYBOARD-MULTI-IOS-013", "Telex a+s → á", "VI active", "Type a,s", "Composer commits á", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-014", "Telex a+f → à", "VI active", "Type a,f", "Composer commits à", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-015", "Telex a+r → ả", "VI active", "Type a,r", "Composer commits ả", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-016", "Telex a+x → ã", "VI active", "Type a,x", "Composer commits ã", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-017", "Telex a+j → ạ", "VI active", "Type a,j", "Composer commits ạ", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-018", "Telex aaj sequence applies tone after circumflex", "VI active", "Type a,a,j", "Composer commits ậ", "TelexComposerTests.swift", "P0", "Tan"),

        # ── Telex composition — real words (4) ───────────────────────────
        ("KEYBOARD-MULTI-IOS-019", 'Type "việt"', "VI active", "Type v,i,e,t,j", "Field shows việt", "TelexComposerTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-020", 'Type "chương"', "VI active", "Type c,h,u,o,w,n,g", "Field shows chương", "TelexComposerTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-021", 'Type "người"', "VI active", "Type n,g,u,o,w,i,f", "Field shows người", "TelexComposerTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-022", 'Type "tiếng"', "VI active", "Type t,i,e,s,n,g", "Field shows tiếng", "TelexComposerTests.swift + manual", "P0", "Tan"),

        # ── Telex backspace (5) ──────────────────────────────────────────
        ("KEYBOARD-MULTI-IOS-023", "Backspace un-composes one step", 'VI active, "việt" composed', "Press ⌫", "Field shows việ (marked text updated via setMarkedText)", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-024", "Backspace from empty buffer falls back to deleteBackward", "VI active, no composing state", "Press ⌫", "textDocumentProxy.deleteBackward called", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-025", "Backspace removes mark before vowel", 'VI active, "â" composed', "Press ⌫", "Field shows a (â un-composed via stripOneLayer)", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-026", "Backspace clears tone before mark", 'VI active, "ậ" composed', "Press ⌫", "Field shows â (tone removed)", "TelexComposerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-027", "Empty-buffer backspace calls unmarkText explicitly", 'VI active, single char "v" composed', "Press ⌫", "textDocumentProxy.unmarkText fires + field clears (no stuck marked region)", "KeyboardViewControllerTests.swift + manual", "P0", "Tan"),

        # ── Telex edge cases (4) ─────────────────────────────────────────
        ("KEYBOARD-MULTI-IOS-028", "Space mid-cluster commits pending via composer", 'VI active, "vie" composing', "Press space", 'Field shows "vie " (vowel cluster committed THEN space appended — not replaced)', "KeyboardViewControllerTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-029", "Language swipe mid-cluster commits + resets", 'VI active, "vie" composing', "Swipe right on space bar", 'Composing region committed as "vie"; composer.reset() called', "KeyboardViewControllerTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-030", "Paste mid-cluster clears composer", 'VI active, "vie" composing', "System paste fires", "Composer reset on next key; pasted text intact", "KeyboardViewControllerTests.swift + manual", "P1", "Tan"),
        ("KEYBOARD-MULTI-IOS-031", "Composing length cap at 32 chars", "VI active", "Type 33 vowel keys", "Composer commits at length 32; new key starts fresh", "TelexComposerTests.swift", "P2", "Tan"),

        # ── Layout swap (7) ──────────────────────────────────────────────
        ("KEYBOARD-MULTI-IOS-032", "EN baseline unchanged", "EN active", 'Type "hello"', "Identical keystroke output to pre-Tier-β iOS keyboard (regression)", "EnglishLanguagePackTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-033", "AZERTY French layout", "FR active", 'Tap QWERTY position of "a"', 'Letter "a" emitted (AZERTY puts a in QWERTY-q slot)', "FrenchLanguagePackTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-034", "Spanish ñ key", "ES active", "Tap ñ key", "Field shows ñ", "SpanishLanguagePackTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-035", "German QWERTZ layout", "DE active", 'Tap QWERTY position of "z"', 'Letter "z" emitted (QWERTZ y/z swapped)', "GermanLanguagePackTests.swift", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-036", "German ß key", "DE active", "Tap ß key", "Field shows ß", "GermanLanguagePackTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-037", "Portuguese ç key", "PT active", "Tap ç key", "Field shows ç", "PortugueseLanguagePackTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-038", "Language cycle order = VI > FR > ES > DE > IT > PT (Tan's preference, matches Android)", "All enabled", "Swipe right on space 6 times", "Cycles through VI, FR, ES, DE, IT, PT, back to start", "KeyboardControllerTests.swift", "P0", "Tan"),

        # ── Swipe-space gesture (3) — REPLACES globe-cycle for iOS ──────
        ("KEYBOARD-MULTI-IOS-039", "Swipe-right on space-bar cycles enabled languages", "VI + EN enabled", "Swipe right on space bar (>80 px)", "activeLanguageId advances; space-bar label updates", "KeyboardViewControllerTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-040", "System iOS globe key opens system keyboard picker (we DO NOT intercept)", "VI active", "Tap iOS system globe key (bottom-left)", "iOS native picker / next system keyboard — DraftRight does NOT handle", "Manual on iPhone", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-041", "Single enabled language: swipe-space is no-op", "Only EN enabled", "Swipe right on space bar", "Nothing happens; tap on space still emits space", "KeyboardViewControllerTests.swift + manual", "P1", "Tan"),

        # ── Long-press accent picker (5) ─────────────────────────────────
        ("KEYBOARD-MULTI-IOS-042", "Long-press a in ES shows á à ä â ã", "ES active", "Hold a key 400ms", "AccentPopupView appears above key with 5 options + a", "AccentPopupViewTests + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-043", "Long-press ñ in ES: no popup", "ES active", "Hold ñ key", "No popup (ñ not in accents map); short-press commits ñ", "SpanishLanguagePackTests.swift + manual", "P1", "Tan"),
        ("KEYBOARD-MULTI-IOS-044", "Long-press e in FR shows é è ê ë variants", "FR active", "Hold e key", "Popup with é è ê ë", "FrenchLanguagePackTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-045", "Finger up without drag = short-press", "ES active", "Hold a, release before drag", "Plain a committed; popup dismisses", "AccentPopupViewTests + manual", "P1", "Tan"),
        ("KEYBOARD-MULTI-IOS-046", "Drag selects + release commits via UIPanGestureRecognizer", "ES active", "Hold a, drag right, release on á", "Field shows á", "Manual", "P0", "Tan"),

        # ── Space-bar label (1) ──────────────────────────────────────────
        ("KEYBOARD-MULTI-IOS-047", "Space bar shows current language displayName", "VI active", "Look at space bar", 'Reads "Tiếng Việt" (or current.displayName)', "Manual on iPhone", "P0", "Tan"),

        # ── Settings persistence (3) ─────────────────────────────────────
        ("KEYBOARD-MULTI-IOS-048", "activeLanguageId survives keyboard re-bind", "VI active", "Switch to another keyboard then back to DraftRight", "VI still active", "KeyboardControllerTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-049", "enabledLanguageIds reorder persists across App Group", "VI > EN order saved in Flutter app", "Reorder to EN > VI in Settings; relaunch keyboard ext", "Swipe-space cycles in new order", "Manual on iPhone", "P1", "Tan"),
        ("KEYBOARD-MULTI-IOS-050", "Disabling all languages forces EN", "User disables all", "Open keyboard", 'EN forced; activeLanguageId="en"; no crash', "KeyboardControllerTests.swift + manual", "P1", "Tan"),

        # ── Backward compatibility (4) ───────────────────────────────────
        ("KEYBOARD-MULTI-IOS-051", "EN-only users see no behavior change", "Fresh install, no setting changes", 'Type "hello world"', "Output identical to pre-Tier-β iOS keyboard", "EnglishLanguagePackTests.swift + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-052", "Existing tone toolbar still functions", "VI active", 'Type text, tap "Polished"', "Tone rewrite flow unchanged; replaceAllText calls unmarkText FIRST", "Manual + integration", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-053", "Share extension unchanged", "EN active", "Long-press text → Share → DraftRight", "Share rewrite flow unchanged", "Manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-054", "/rewrite payload includes only text + tone (no inputLanguage)", "VI active", "Tap Polished", "Backend request body has text + tone only", "BackendClient inspection", "P0", "Tan"),

        # ── iOS-specific tests (2) ───────────────────────────────────────
        ("KEYBOARD-MULTI-IOS-055", "Full Access disabled → typing works, rewrite shows banner", "Allow Full Access OFF for DraftRight Keyboard in iOS Settings", "Open keyboard, type, tap tone button", "Typing works locally (Telex composer is offline). Rewrite shows 'Allow Full Access in Settings' banner — no network call attempted.", "Manual on iPhone 14/SE", "P0", "Tan"),
        ("KEYBOARD-MULTI-IOS-056", "Memory under 50 MB with all 7 packs loaded", "All 7 enabledLanguageIds", "Type one sentence in each language; observe Xcode Allocations", "Peak ext memory < 50 MB (Apple cap ≈ 70 MB)", "Xcode Instruments + manual", "P0", "Tan"),
    ]
    for r in rows:
        ws.append(r)
        for c in range(1, len(r) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 26, "B": 56, "C": 36, "D": 50, "E": 60, "F": 34, "G": 10, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 50
    wb.save(XLSX)
    print(f"Added {SHEET} with {len(rows)} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
