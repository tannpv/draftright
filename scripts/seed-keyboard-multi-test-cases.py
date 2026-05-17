#!/usr/bin/env python3
"""Append KEYBOARD-MULTI test cases to docs/test-cases.xlsx — 54 rows.

Idempotent guard: aborts if the sheet already exists. Drop the sheet manually
(or delete it via openpyxl) before re-running."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX = Path(__file__).resolve().parents[1] / "docs" / "test-cases.xlsx"
SHEET = "KEYBOARD-MULTI"
HEADERS = ["TC-ID", "Title", "Preconditions", "Steps", "Expected", "Verified by", "Priority", "Owner"]


def main() -> int:
    wb = load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        print(f"Sheet {SHEET} exists — aborting")
        return 1
    ws = wb.create_sheet(SHEET)
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    rows = [
        # ── Telex composition — core rules (12) ─────────────────────────
        ("KEYBOARD-MULTI-001", "Telex aa → â", "VI active", "Type a,a", "Composer commits â", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-002", "Telex oo → ô", "VI active", "Type o,o", "Composer commits ô", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-003", "Telex ee → ê", "VI active", "Type e,e", "Composer commits ê", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-004", "Telex ow → ơ", "VI active", "Type o,w", "Composer commits ơ", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-005", "Telex uw → ư", "VI active", "Type u,w", "Composer commits ư", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-006", "Telex aw → ă", "VI active", "Type a,w", "Composer commits ă", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-007", "Telex dd → đ", "VI active", "Type d,d", "Composer commits đ", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-008", "Telex aaj → ậ", "VI active", "Type a,a,j", "Composer commits ậ (â + dot)", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-009", "Telex uow → ươ", "VI active", "Type u,o,w", "Composer commits ươ", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-010", "Telex uowj → ượ", "VI active", "Type u,o,w,j", "Composer commits ượ", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-011", "Telex case preserved on shift", "VI active, shift on", "Type A,A", "Composer commits Â", "TelexComposerTest.kt", "P1", "Tan"),
        ("KEYBOARD-MULTI-012", "Telex no-op for non-rule keys", "VI active", "Type q", "Direct commit q (no composing)", "TelexComposerTest.kt", "P0", "Tan"),

        # ── Telex composition — tones (6) ────────────────────────────────
        ("KEYBOARD-MULTI-013", "Telex a+s → á", "VI active", "Type a,s", "Composer commits á", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-014", "Telex a+f → à", "VI active", "Type a,f", "Composer commits à", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-015", "Telex a+r → ả", "VI active", "Type a,r", "Composer commits ả", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-016", "Telex a+x → ã", "VI active", "Type a,x", "Composer commits ã", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-017", "Telex a+j → ạ", "VI active", "Type a,j", "Composer commits ạ", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-018", "Telex aaj sequence applies tone after circumflex", "VI active", "Type a,a,j", "Composer commits ậ", "TelexComposerTest.kt", "P0", "Tan"),

        # ── Telex composition — real words (4) ───────────────────────────
        ("KEYBOARD-MULTI-019", 'Type "việt"', "VI active", "Type v,i,e,t,j", "Field shows việt", "TelexComposerTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-020", 'Type "chương"', "VI active", "Type c,h,u,o,w,n,g", "Field shows chương", "TelexComposerTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-021", 'Type "người"', "VI active", "Type n,g,u,o,w,i,f", "Field shows người", "TelexComposerTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-022", 'Type "tiếng"', "VI active", "Type t,i,e,s,n,g", "Field shows tiếng", "TelexComposerTest.kt + manual", "P0", "Tan"),

        # ── Telex backspace (5) ──────────────────────────────────────────
        ("KEYBOARD-MULTI-023", "Backspace un-composes one step", 'VI active, "việt" composed', "Press ⌫", "Field shows việ (last char un-composed via composer)", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-024", "Backspace at empty composer passes through", "VI active, no composing state", "Press ⌫", "InputConnection.deleteSurroundingText called", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-025", "Backspace removes mark before vowel", 'VI active, "â" composed', "Press ⌫", "Field shows a (â un-composed)", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-026", "Backspace clears tone before mark", 'VI active, "ậ" composed', "Press ⌫", "Field shows â (tone removed)", "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-027", "Multiple backspaces clear composer state", 'VI active, "việt" composed', "Press ⌫ four times", "Composer state empty; further ⌫ deletes via IC", "TelexComposerTest.kt", "P0", "Tan"),

        # ── Telex edge cases (4) ─────────────────────────────────────────
        ("KEYBOARD-MULTI-028", "Non-letter mid-cluster commits pending", 'VI active, "vie" composing', "Type space", 'Field shows "vie " (vowel cluster committed)', "TelexComposerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-029", "Language switch mid-cluster commits + resets", 'VI active, "vie" composing', "Tap globe to switch to EN", 'Composing region committed as "vie"; composer.reset() called', "KeyboardControllerTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-030", "Paste mid-cluster clears composer", 'VI active, "vie" composing', "System paste fires commitText", "Composer onStartInput-like reset; pasted text intact", "KeyboardControllerTest.kt + manual", "P1", "Tan"),
        ("KEYBOARD-MULTI-031", "Composing length cap at 32 chars", "VI active", "Type 33 vowel keys", "Composer commits at length 32; new key starts fresh", "TelexComposerTest.kt", "P2", "Tan"),

        # ── Layout swap (7) ──────────────────────────────────────────────
        ("KEYBOARD-MULTI-032", "EN baseline unchanged", "EN active", 'Type "hello"', "Identical keystroke output to today (regression)", "EnglishLanguagePackTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-033", "AZERTY French layout", "FR active", 'Tap QWERTY position of "a"', 'Letter "a" emitted (AZERTY puts a in QWERTY-q slot)', "FrenchLanguagePackTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-034", "Spanish ñ key", "ES active", "Tap ñ key", "Field shows ñ", "SpanishLanguagePackTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-035", "German QWERTZ layout", "DE active", 'Tap QWERTY position of "z"', 'Letter "z" emitted (QWERTZ y/z swapped)', "GermanLanguagePackTest.kt", "P0", "Tan"),
        ("KEYBOARD-MULTI-036", "German ß key", "DE active", "Tap ß key", "Field shows ß", "GermanLanguagePackTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-037", "Portuguese ç key", "PT active", "Tap ç key", "Field shows ç", "PortugueseLanguagePackTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-038", "Language cycle order = VI > FR > ES > DE > IT > PT", "All enabled", "Tap globe 6 times", "Cycles through VI, FR, ES, DE, IT, PT, back to start", "KeyboardControllerTest.kt", "P0", "Tan"),

        # ── Globe key (3) ────────────────────────────────────────────────
        ("KEYBOARD-MULTI-039", "Globe tap cycles enabled languages", "VI + EN enabled", "Tap globe", "activeLanguageId flips between VI and EN", "KeyboardControllerTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-040", "Globe long-press opens IME picker", "VI active", "Long-press globe 500ms", "InputMethodManager.showInputMethodPicker() called", "KeyboardControllerTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-041", "Single enabled language: globe tap is no-op", "Only EN enabled", "Tap globe", "activeLanguageId unchanged; strip not shown", "KeyboardControllerTest.kt + manual", "P1", "Tan"),

        # ── Long-press accent picker (5) ─────────────────────────────────
        ("KEYBOARD-MULTI-042", "Long-press a in ES shows á à ä â ã", "ES active", "Hold a key 300ms", "AccentPopupView appears above key with 5 options + a", "AccentPopupViewTest + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-043", "Long-press ñ in ES: no popup", "ES active", "Hold ñ key", "No popup (ñ not in accents map); short-press commits ñ", "SpanishLanguagePackTest.kt + manual", "P1", "Tan"),
        ("KEYBOARD-MULTI-044", "Long-press a in FR shows é è ê variants for e key", "FR active", "Hold e key", "Popup with é è ê ë", "FrenchLanguagePackTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-045", "Finger up without drag = short-press", "ES active", "Hold a, release before drag", "Plain a committed; popup dismisses", "AccentPopupViewTest + manual", "P1", "Tan"),
        ("KEYBOARD-MULTI-046", "Drag selects + release commits", "ES active", "Hold a, drag right, release on á", "Field shows á", "manual", "P0", "Tan"),

        # ── Space-bar label (1) ──────────────────────────────────────────
        ("KEYBOARD-MULTI-047", "Space bar shows current language displayName", "VI active", "Look at space bar", 'Reads "Tiếng Việt" (or current.displayName)', "manual on emulator", "P0", "Tan"),

        # ── Settings persistence (3) ─────────────────────────────────────
        ("KEYBOARD-MULTI-048", "activeLanguageId survives Force-Stop", "VI active", "Force-stop app; relaunch IME", "VI still active", "KeyboardControllerTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-049", "enabledLanguageIds reorder persists", "VI > EN order saved", "Reorder to EN > VI in Settings; relaunch", "Globe cycles EN > VI in new order", "manual on emulator", "P1", "Tan"),
        ("KEYBOARD-MULTI-050", "Disabling all languages forces EN", "User disables all", "Open keyboard", 'EN forced; activeLanguageId="en"; toast/log warning', "KeyboardControllerTest.kt + manual", "P1", "Tan"),

        # ── Backward compatibility (4) ───────────────────────────────────
        ("KEYBOARD-MULTI-051", "EN-only users see no behavior change", "Fresh install, no setting changes", 'Type "hello world"', "Output identical to pre-Tier-β", "EnglishLanguagePackTest.kt + manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-052", "Existing tone toolbar still functions", "VI active", 'Type text, tap "Polished"', "Tone rewrite flow unchanged", "manual + integration", "P0", "Tan"),
        ("KEYBOARD-MULTI-053", "Share intent unchanged", "EN active", "Long-press text → Share → DraftRight", "Share rewrite flow unchanged", "manual", "P0", "Tan"),
        ("KEYBOARD-MULTI-054", "/rewrite payload includes only text + tone (no inputLanguage)", "VI active", "Tap Polished", "Backend request body has text + tone only", "BackendClient inspection", "P0", "Tan"),
    ]
    for r in rows:
        ws.append(r)
        for c in range(1, len(r) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 22, "B": 52, "C": 28, "D": 50, "E": 52, "F": 32, "G": 10, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 50
    wb.save(XLSX)
    print(f"Added {SHEET} with {len(rows)} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
