#!/usr/bin/env python3
"""Append SINGLE-INST test cases to docs/test-cases.xlsx — 6 rows.

Single-instance enforcement for the Windows app (2.2.8). Idempotent
guard: aborts if the sheet already exists."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX = Path(__file__).resolve().parents[1] / "docs" / "test-cases.xlsx"
SHEET = "SINGLE-INST"
HEADERS = ["TC-ID", "Title", "Preconditions", "Steps", "Expected", "Verified by", "Priority", "Owner"]


def main() -> int:
    wb = load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        print(f"Sheet {SHEET} exists — aborting")
        return 1
    ws = wb.create_sheet(SHEET)
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    rows = [
        ("SINGLE-INST-001", "First launch starts normally",
         "No DraftRight.exe running",
         "Launch DraftRight.exe",
         "Single tray icon appears; single DraftRight.exe in Task Manager",
         "Manual on Windows 11 x64", "P0", "Tan"),
        ("SINGLE-INST-002", "Second launch exits silently",
         "One DraftRight.exe already running with tray icon",
         "Launch DraftRight.exe a second time",
         "Second process exits immediately; still exactly one tray icon; still one DraftRight.exe in Task Manager",
         "Manual on Windows 11 x64", "P0", "Tan"),
        ("SINGLE-INST-003", "No duplicate taskbar icon when RewritePanel is open",
         "DraftRight running, rewrite panel open via hotkey",
         "Launch DraftRight.exe a second time",
         "Taskbar still shows exactly one DraftRight entry; no flicker",
         "Manual on Windows 11 x64", "P0", "Tan"),
        ("SINGLE-INST-004", "Mutex released on graceful exit",
         "DraftRight running",
         "Quit via tray menu → Exit; wait 2s; launch DraftRight.exe again",
         "New process launches normally; tray icon reappears",
         "Manual on Windows 11 x64", "P0", "Tan"),
        ("SINGLE-INST-005", "Mutex released on process kill",
         "DraftRight running",
         "Task Manager → End Task on DraftRight.exe; wait 2s; launch DraftRight.exe",
         "New process launches normally (kernel reclaims the abandoned mutex)",
         "Manual on Windows 11 x64", "P1", "Tan"),
        ("SINGLE-INST-006", "Different Windows users can each run their own instance",
         "User A signed in with DraftRight running",
         "Fast-switch to User B; launch DraftRight.exe as User B",
         "User B's instance runs (Mutex is Local\\, not Global\\); both users have their own tray icon",
         "Manual on Windows 11 x64 multi-user host", "P2", "Tan"),
    ]
    for r in rows:
        ws.append(r)
        for c in range(1, len(r) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 22, "B": 52, "C": 36, "D": 50, "E": 60, "F": 32, "G": 10, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 50
    wb.save(XLSX)
    print(f"Added {SHEET} with {len(rows)} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
