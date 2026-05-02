#!/usr/bin/env python3
"""Seed docs/test-cases.xlsx with the EXTTOK sheet for the extension-tokens
implementation plan. Idempotent: re-running will overwrite the EXTTOK sheet
without touching other sheets."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

XLSX = Path(__file__).resolve().parents[1] / "docs" / "test-cases.xlsx"
SHEET = "EXTTOK"
HEADERS = ["TC-ID", "Title", "Preconditions", "Steps", "Expected", "Verified by", "Priority", "Owner"]

ROWS = [
    ("EXTTOK-001", "Backend mints token with dr_ext_ prefix",
     "user authenticated with session JWT",
     "POST /auth/extension-tokens with valid device_id, device_name",
     "200, body has token starting with dr_ext_ and id (UUID)",
     "extension-token.service.spec.ts: 'mint returns plaintext token with dr_ext_ prefix'", "P0", "Tan"),
    ("EXTTOK-002", "Backend stores only sha256(token), never plaintext",
     "fresh DB",
     "Mint a token, then SELECT token_hash FROM extension_tokens",
     "token_hash column is 64-char hex (sha256). No plaintext anywhere in DB.",
     "extension-token.service.spec.ts: 'stores only the sha256 hash'", "P0", "Tan"),
    ("EXTTOK-003", "Re-mint for same (user, device_id) revokes the old row",
     "user has one active extension token for device D",
     "Mint again with same device_id",
     "Old row has revoked_at set. New row is the only active row.",
     "extension-token.service.spec.ts: 'revokes the existing active token...' + matrix A8", "P0", "Tan"),
    ("EXTTOK-004", "Validate returns user_id + scopes for active token",
     "active token T exists for user U",
     "Call ExtensionTokenService.validate(T)",
     "Returns { tokenId, userId: U, scopes: ['rewrite'] }",
     "extension-token.service.spec.ts: 'returns user_id and scopes for valid token'", "P0", "Tan"),
    ("EXTTOK-005", "Validate returns null for revoked token",
     "token T was revoked",
     "Call ExtensionTokenService.validate(T)",
     "Returns null. Token is unusable.",
     "extension-token.service.spec.ts + matrix A7", "P0", "Tan"),
    ("EXTTOK-006", "Validate returns null for token without dr_ext_ prefix",
     "any string without dr_ext_ prefix",
     "Call validate('not-an-extension-token')",
     "Returns null without DB lookup.",
     "extension-token.service.spec.ts: 'returns null for token without dr_ext_ prefix'", "P0", "Tan"),
    ("EXTTOK-007", "Revoke endpoint sets revoked_at for matching (user_id, id) only",
     "user U has token T1 (id i1); user V has token T2 (id i2)",
     "DELETE /auth/extension-tokens/i1 with U's JWT",
     "204. T1 revoked. T2 untouched. DELETE /auth/extension-tokens/i2 with U's JWT does NOT revoke T2.",
     "extension-token.service.spec.ts + extension-token.controller.spec.ts", "P0", "Tan"),
    ("EXTTOK-008", "POST /rewrite accepts user JWT (back-compat)",
     "user authenticated with session JWT",
     "POST /rewrite with Authorization: Bearer <jwt>",
     "200 with rewritten_text",
     "rewrite-auth.guard.spec.ts + matrix A1", "P0", "Tan"),
    ("EXTTOK-009", "POST /rewrite accepts extension token with rewrite scope",
     "user has active extension token T with scope rewrite",
     "POST /rewrite with Authorization: Bearer <T>",
     "200 with rewritten_text",
     "rewrite-auth.guard.spec.ts + matrix A3", "P0", "Tan"),
    ("EXTTOK-010", "Non-rewrite endpoints reject extension tokens",
     "valid extension token T",
     "GET /auth/me with Authorization: Bearer <T>",
     "401 (extension tokens have only 'rewrite' scope; /auth/me requires JWT)",
     "rewrite-auth.guard usage check + matrix A4", "P0", "Tan"),
    ("EXTTOK-011", "Flutter mints on login (calls POST /auth/extension-tokens)",
     "fresh install, no token in shared storage",
     "Login in main app",
     "ExtensionTokenService.ensureMinted is called; token persists in App Group keychain (iOS) or SharedPreferences (Android)",
     "matrix B1, B2, C1, C2", "P0", "Tan"),
    ("EXTTOK-012", "Flutter clears stored token on logout",
     "user logged in with extension token in shared storage",
     "Tap Logout in main app",
     "Token removed from keychain / SharedPreferences. Extension calls 401.",
     "matrix B6, C5", "P1", "Tan"),
    ("EXTTOK-013", "Flutter device_id is generated once and persisted",
     "no prior install state",
     "Call ExtensionTokenService.deviceId() twice",
     "Returns the same UUIDv4 both times. Persisted to SharedPreferences as draftright.deviceId.",
     "extension_token_service_test.dart: 'deviceId is generated once and persisted'", "P1", "Tan"),
    ("EXTTOK-014", "iOS keychain item readable from all 3 targets via shared access-group",
     "main app has written extension token to App Group keychain",
     "Open keyboard extension; open share extension. Both call BackendClient.rewrite().",
     "Both reads succeed. Both API calls authenticate.",
     "matrix B2, B7 (cross-target read confirmed implicitly)", "P0", "Tan"),
    ("EXTTOK-015", "iOS keyboard rewrite still works 30 min after main app last used",
     "main app idle for 30+ min, but extension token is in keychain",
     "Open Notes, switch to DraftRight keyboard, request a rewrite",
     "200 result. NO 'Please login' error. (This is the actual bug we're fixing.)",
     "matrix B3", "P0", "Tan"),
    ("EXTTOK-016", "iOS share extension rewrite still works 30 min after main app last used",
     "main app idle for 30+ min",
     "In Safari, share text -> DraftRight Action -> pick tone",
     "200 result. NO 'Please login' error.",
     "matrix B4", "P0", "Tan"),
    ("EXTTOK-017", "Android IME rewrite still works 30 min after main app last used",
     "main app idle for 30+ min",
     "Open any Android app, switch to DraftRight IME, tap a tone in toolbar",
     "200 result. NO 'Please login' error.",
     "matrix C3", "P0", "Tan"),
    ("EXTTOK-018", "Upgrade-in-place: old access-JWT works until first main-app launch mints new token",
     "user is on prior build with access-JWT in shared storage; upgrades to new build without opening main app",
     "Use keyboard within access-JWT TTL",
     "Rewrite succeeds via access-JWT fallback path. Then opening main app once mints the new token; subsequent calls use it.",
     "matrix D1-D5", "P1", "Tan"),
    ("EXTTOK-019", "Re-mint for same device rotates token (old presents as 401)",
     "device D has active token T1",
     "Re-login on same device (mints T2). Try to use T1.",
     "T1 is revoked. T2 is the only active token. T1 in any HTTP request returns 401.",
     "matrix A8", "P1", "Tan"),
    ("EXTTOK-020", "After logout, server sees presented token as revoked / unrecognized",
     "user logged in with extension token T",
     "Logout in main app. Then attempt rewrite from extension via stale T.",
     "T returns 401. extension_tokens row reflects revoked_at is set OR has been deleted entirely.",
     "matrix B6 + manual psql check", "P1", "Tan"),
]

XLSX.parent.mkdir(parents=True, exist_ok=True)
if XLSX.exists():
    wb = load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]

ws = wb.create_sheet(SHEET)
ws.append(HEADERS)
for row in ROWS:
    ws.append(row)

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 50
ws.column_dimensions["E"].width = 50
ws.column_dimensions["F"].width = 50
ws.column_dimensions["G"].width = 8
ws.column_dimensions["H"].width = 8

wb.save(XLSX)
print(f"Wrote {XLSX} with {len(ROWS)} rows on sheet '{SHEET}'.")
